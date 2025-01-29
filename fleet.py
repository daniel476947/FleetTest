from flask import Flask, request, jsonify, render_template, send_file, redirect, url_for
from flask_cors import CORS
from pymongo import MongoClient
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bson.objectid import ObjectId
import os
import re
import tempfile
import logging
import gridfs  # Import GridFS
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)

# Configure logging
logging.basicConfig(level=logging.INFO, filename='app.log',
                    format='%(asctime)s %(levelname)s:%(message)s')

# MongoDB Atlas connection string
MONGO_URI = os.getenv('MONGO_URI')

if not MONGO_URI:
    logging.error("MONGO_URI environment variable not set.")
    raise EnvironmentError("MONGO_URI environment variable not set.")

# Initialize the client
client = MongoClient(MONGO_URI)

# Specify the database and collection
db = client['Fleettest']
collection = db['Fleet']

# Initialize GridFS
fs = gridfs.GridFS(db)

REG_NO_PATTERN = r'^[A-Z]{3} [0-9]{1,4}$'  # 3 uppercase letters, space, 1-4 digits

def build_query(form_data):
    """
    Builds a MongoDB query dictionary based on form data.
    """
    query = {}
    search_term = form_data.get('search', '').strip()
    
    if search_term:
        search_words = search_term.split()
        query["$and"] = []
        for word in search_words:
            or_conditions = [
                {"Registration No": {"$regex": word, "$options": "i"}},
                {"Make": {"$regex": word, "$options": "i"}},
                {"Model": {"$regex": word, "$options": "i"}},
                {"Chassis No": {"$regex": word, "$options": "i"}},
            ]
            if word.isdigit():
                or_conditions.append({"Year": int(word)})
            query["$and"].append({"$or": or_conditions})
    
    # Apply filters
    filters = {
        "Vehicle Type": form_data.get('vehicle_type'),
        "Main Colour": form_data.get('main_colour'),
        "Secondary Colour": form_data.get('secondary_colour'),
        "Fuel": form_data.get('fuel'),
        "Status": form_data.get('status'),
        "Location": form_data.get('location')
    }
    
    for key, value in filters.items():
        if value:
            query[key] = value
    
    # Registration Status
    registration_status = form_data.get('registration_status')
    if registration_status == 'Registered':
        query["Registration No"] = {"$nin": [None, ""]}
    elif registration_status == 'Unregistered':
        query["Registration No"] = {"$in": [None, ""]}
    
    return query

def allowed_file(filename):
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

@app.route('/', methods=['GET', 'POST'])
def view_fleet():
    try:
        if request.method == 'POST':
            form_data = request.form
            # For POST requests, reset to the first page
            page = 1
        else:
            form_data = request.args  # To handle GET requests if needed
            # Get the current page from query parameters, default to 1
            page = int(request.args.get('page', 1))
        
        # Number of records per page
        PER_PAGE = 10
        
        # Build the query based on form data
        query = build_query(form_data if request.method == 'POST' else {})
        
        # Count total matching documents
        total_documents = collection.count_documents(query)
        
        # Calculate total pages
        total_pages = (total_documents + PER_PAGE - 1) // PER_PAGE  # Ceiling division
        
        # Ensure the current page is within valid range
        if page < 1:
            page = 1
        elif page > total_pages and total_pages != 0:
            page = total_pages
        
        # Fetch the documents for the current page
        vehicles_cursor = collection.find(query).skip((page - 1) * PER_PAGE).limit(PER_PAGE)
        vehicles = list(vehicles_cursor)
        
        for vehicle in vehicles:
            vehicle['_id'] = str(vehicle['_id'])  # Convert ObjectId to string for rendering
        
        # Prepare pagination data
        pagination = {
            'total_pages': total_pages,
            'current_page': page,
            'has_prev': page > 1,
            'has_next': page < total_pages,
            'prev_page': page - 1,
            'next_page': page + 1,
            'pages': list(range(1, total_pages + 1))
        }
        
        return render_template('viewfleet.html', vehicles=vehicles, pagination=pagination, form_data=form_data)
    
    except Exception as e:
        logging.error(f"Error fetching data: {e}")
        return jsonify({"success": False, "message": "Error fetching data."}), 500


@app.route('/upload', methods=['POST'])
def upload():
    try:
        # 1. Grab the raw Registration No from form, strip whitespace
        reg_no_raw = request.form.get('Registration No', '').strip()

        # 2. Convert to uppercase
        reg_no = reg_no_raw.upper()

        # 3. Validate with regex
        if reg_no and not re.match(REG_NO_PATTERN, reg_no):
            logging.warning(f"Invalid Registration No format: {reg_no}")
            return jsonify({"success": False, "message": "Invalid Registration No format."}), 400

        # Convert Year, Capacity to int
        year_str = request.form.get('Year', '').strip()
        capacity_str = request.form.get('Capacity', '').strip()
        
        if not year_str.isdigit() or not capacity_str.isdigit():
            logging.warning("Year and/or Capacity are not valid integers.")
            return jsonify({"success": False, "message": "Year and Capacity must be valid integers."}), 400

        year_val = int(year_str)        
        capacity_val = int(capacity_str)

        # Handle Secondary Colour
        secondary_colour = request.form.get('Secondary Colour', '').strip()
        if secondary_colour.lower() == 'none':
            secondary_colour = ""

        # Build vehicle_data dictionary
        vehicle_data = {
            "Registration No": reg_no,
            "Make": request.form.get('Make', '').strip(),
            "Model": request.form.get('Model', '').strip(),
            "Vehicle Type": request.form.get('Vehicle Type', '').strip(),
            "Year": year_val,
            "Main Colour": request.form.get('Main Colour', '').strip(),
            "Secondary Colour": secondary_colour,
            "Fuel": request.form.get('Fuel', '').strip(),
            "Capacity": capacity_val,
            "Chassis No": request.form.get('Chassis No', '').strip(),
            "Model No": request.form.get('Model No', '').strip(),
            "Status": request.form.get('Status', '').strip(),
            "Location": request.form.get('Location', '').strip(),
            "image_ids": []  # Initialize empty list for image IDs
        }

        # Handle Image Uploads
        if 'images' in request.files:
            images = request.files.getlist('images')
            for image in images:
                # Only process files with a non-empty filename
                if image and image.filename and allowed_file(image.filename):
                    filename = secure_filename(image.filename)
                    # Save image to GridFS
                    image_id = fs.put(image.read(), filename=filename, content_type=image.content_type)
                    vehicle_data['image_ids'].append(str(image_id))
                elif image and image.filename:
                    # Invalid file type
                    logging.warning(f"Invalid file type for image: {image.filename}")
                    return jsonify({"success": False, "message": f"Invalid file type for image: {image.filename}"}), 400
                # If image.filename is empty, skip processing

        # 4. Insert into MongoDB
        result = collection.insert_one(vehicle_data)
        logging.info(f"Data successfully inserted with ID: {result.inserted_id}")
        return jsonify({"success": True, "message": "Data uploaded successfully!"}), 200

    except ValueError:
        # If int() conversions fail
        logging.error("Year and Capacity must be valid integers.")
        return jsonify({"success": False, "message": "Year and Capacity must be valid integers."}), 400
    except Exception as e:
        logging.error(f"Error inserting data: {e}")
        return jsonify({"success": False, "message": "Error uploading data."}), 500


@app.route('/view/<id>', methods=['GET'])
def view_vehicle(id):
    try:
        # Fetch the vehicle document by its ObjectId
        vehicle = collection.find_one({"_id": ObjectId(id)})
        if not vehicle:
            logging.warning(f"Vehicle not found with ID: {id}")
            return jsonify({"success": False, "message": "Vehicle not found."}), 404

        # Convert ObjectId to string for rendering
        vehicle['_id'] = str(vehicle['_id'])

        return render_template('view_vehicle.html', vehicle=vehicle)
    except Exception as e:
        logging.error(f"Error viewing vehicle: {e}")
        return jsonify({"success": False, "message": "Error viewing vehicle."}), 500


@app.route('/image/<image_id>')
def get_image(image_id):
    try:
        # Validate and convert the image_id to ObjectId
        obj_id = ObjectId(image_id)
    except Exception as e:
        app.logger.error(f"Invalid image_id format: {image_id} - {e}")
        return jsonify({"success": False, "message": "Invalid image ID."}), 400

    try:
        # Retrieve the image from GridFS
        image = fs.get(obj_id)
        return send_file(
            image,
            mimetype=image.content_type,
            as_attachment=False,
            download_name=image.filename  # Updated to 'download_name' for Flask 2.0+
        )
    except gridfs.NoFile:
        app.logger.error(f"No file found with image_id: {image_id}")
        return jsonify({"success": False, "message": "Image not found."}), 404
    except Exception as e:
        app.logger.error(f"Error retrieving image {image_id}: {e}")
        return jsonify({"success": False, "message": "Error retrieving image."}), 500


@app.route('/export_fleet', methods=['GET'])
def export_fleet():
    try:
        query = build_query(request.args)
        vehicles = list(collection.find(query))
        for vehicle in vehicles:
            vehicle.pop('_id', None)  # Safely remove _id if it exists

        if not vehicles:
            logging.info("No data to export.")
            return jsonify({"success": False, "message": "No data to export."}), 400

        df = pd.DataFrame(vehicles)

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            excel_file = tmp.name
            df.to_excel(excel_file, index=False)

        # Load the workbook and sheet to apply formatting
        wb = load_workbook(excel_file)
        ws = wb.active

        # Apply right alignment to all cells except headers
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='right')

        # Save the modified Excel file
        wb.save(excel_file)

        logging.info(f"Exported data to Excel file: {excel_file}")
        return send_file(excel_file, as_attachment=True, download_name='fleet_data.xlsx')
    except Exception as e:
        logging.error(f"Error exporting data: {e}")
        return jsonify({"success": False, "message": "Error exporting data."}), 500


@app.route('/edit/<id>', methods=['GET', 'POST'])
def edit_vehicle(id):
    try:
        vehicle = collection.find_one({"_id": ObjectId(id)})
        if not vehicle:
            logging.warning(f"Vehicle not found with ID: {id}")
            return render_template('editvehicle.html', vehicle={}, error_message="Vehicle not found."), 404

        if request.method == 'GET':
            vehicle['_id'] = str(vehicle['_id'])  # Convert ObjectId to string
            return render_template('editvehicle.html', vehicle=vehicle)

        if request.method == 'POST':
            # Registration No
            reg_no_raw = request.form.get('Registration No', '').strip()
            reg_no = reg_no_raw.upper()
            # Validate format
            if reg_no and not re.match(REG_NO_PATTERN, reg_no):
                logging.warning(f"Invalid Registration No format during edit: {reg_no}")
                error_message = "Invalid Registration No format. Use 3 letters, a space, and 1-4 digits (e.g., PGG 1567)."
                vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
                return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 400

            # Convert Year, Capacity to integers
            year_str = request.form.get('Year', '').strip()
            capacity_str = request.form.get('Capacity', '').strip()
            if not year_str.isdigit() or not capacity_str.isdigit():
                logging.warning("Year and/or Capacity are not valid integers during edit.")
                error_message = "Year and Capacity must be valid integers."
                vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
                return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 400

            year_val = int(year_str)
            capacity_val = int(capacity_str)

            # Handle Secondary Colour
            secondary_colour = request.form.get('Secondary Colour', '').strip()
            if secondary_colour.lower() == 'none':
                secondary_colour = ""

            # Start with existing image_ids
            updated_image_ids = vehicle.get('image_ids', [])

            # Handle Image Deletions (Existing Images)
            images_to_delete = request.form.getlist('delete_images')  # List of image IDs to delete
            if images_to_delete:
                for img_id in images_to_delete:
                    try:
                        fs.delete(ObjectId(img_id))
                        logging.info(f"Deleted image {img_id} associated with vehicle {id}.")
                        # Remove img_id from updated_image_ids
                        if img_id in updated_image_ids:
                            updated_image_ids.remove(img_id)
                    except gridfs.NoFile:
                        logging.warning(f"Image {img_id} not found in GridFS.")
                        error_message = f"Image with ID {img_id} not found."
                        vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
                        return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 404
                    except Exception as e:
                        logging.error(f"Error deleting image {img_id}: {e}")
                        error_message = "An error occurred while deleting images."
                        vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
                        return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 500

            # Handle Image Uploads (New Images)
            if 'images' in request.files:
                images = request.files.getlist('images')
                for image in images:
                    if image and image.filename and allowed_file(image.filename):
                        filename = secure_filename(image.filename)
                        # Save image to GridFS
                        image_id = fs.put(image.read(), filename=filename, content_type=image.content_type)
                        updated_image_ids.append(str(image_id))
                    elif image and image.filename:
                        # Image has a filename but invalid type
                        logging.warning(f"Invalid file type for image during edit: {image.filename}")
                        error_message = f"Invalid file type for image: {image.filename}. Allowed types: png, jpg, jpeg, gif."
                        vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
                        return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 400
                # If image.filename is empty, skip processing

            # Build updated_data dictionary
            updated_data = {
                "Registration No": reg_no,
                "Make": request.form.get('Make', '').strip(),
                "Model": request.form.get('Model', '').strip(),
                "Vehicle Type": request.form.get('Vehicle Type', '').strip(),
                "Year": year_val,
                "Main Colour": request.form.get('Main Colour', '').strip(),
                "Secondary Colour": secondary_colour,
                "Fuel": request.form.get('Fuel', '').strip(),
                "Capacity": capacity_val,
                "Chassis No": request.form.get('Chassis No', '').strip(),
                "Model No": request.form.get('Model No', '').strip(),
                "Status": request.form.get('Status', '').strip(),
                "Location": request.form.get('Location', '').strip(),
                "image_ids": updated_image_ids  # Set to the updated list
            }

            # Update the vehicle in MongoDB
            result = collection.update_one({"_id": ObjectId(id)}, {"$set": updated_data})
            if result.modified_count == 1:
                logging.info(f"Vehicle with ID {id} updated successfully.")
            else:
                logging.info(f"No changes made to vehicle with ID {id}.")

            return redirect(url_for('view_fleet'))

    except ValueError:
        logging.error("Year and Capacity must be valid integers during edit.")
        error_message = "Year and Capacity must be valid integers."
        vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
        return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 400
    except Exception as e:
        logging.error(f"Error editing vehicle: {e}")
        error_message = "An unexpected error occurred while editing the vehicle."
        vehicle['_id'] = str(vehicle['_id'])  # Ensure ID is string
        return render_template('editvehicle.html', vehicle=vehicle, error_message=error_message), 500


@app.route('/delete/<id>', methods=['POST'])
def delete_vehicle(id):
    try:
        # First, fetch the vehicle document to get associated image_ids
        vehicle = collection.find_one({"_id": ObjectId(id)})
        if not vehicle:
            logging.warning(f"Vehicle not found for deletion with ID: {id}")
            return jsonify({"success": False, "message": "Vehicle not found."}), 404

        # Delete associated images from GridFS
        if 'image_ids' in vehicle:
            for img_id in vehicle['image_ids']:
                try:
                    fs.delete(ObjectId(img_id))
                    logging.info(f"Deleted image {img_id} associated with vehicle {id}.")
                except gridfs.NoFile:
                    logging.warning(f"Image {img_id} not found in GridFS.")
                except Exception as e:
                    logging.error(f"Error deleting image {img_id}: {e}")
                    return jsonify({"success": False, "message": "Error deleting images."}), 500

        # Delete the vehicle document
        result = collection.delete_one({"_id": ObjectId(id)})
        if result.deleted_count == 1:
            logging.info(f"Vehicle with ID {id} deleted successfully.")
            return redirect(url_for('view_fleet'))
        else:
            logging.warning(f"Vehicle not found for deletion with ID: {id}")
            return jsonify({"success": False, "message": "Vehicle not found."}), 404
    except Exception as e:
        logging.error(f"Error deleting vehicle: {e}")
        return jsonify({"success": False, "message": "Error deleting vehicle."}), 500    


@app.route('/form')
def index():
    return render_template('form.html')  # Renders the HTML form


if __name__ == '__main__':
    app.run(debug=True)  # Consider setting debug=False in production
